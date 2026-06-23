import logging
import pytz
from datetime import datetime, timedelta, date, time
import django.utils.timezone
from django.db.models import Q
from common.models import UserInOut, CompanyEmployee, CompanyDetails, Locations, WeeklyOff, HolidayTemplates, HolidayTemplateDetails
from common.service import CommonService

logger = logging.getLogger(__name__)

class UserInOutService:
    def __init__(self):
        self.common_service = CommonService()

    def apply_cell_style(self, cell, is_bold=False, has_borders=False, is_centered=False, is_vertically_centered=False, font_size=11):
        from openpyxl.styles import Font, Alignment, Border, Side
        font = Font(name="Calibri", size=font_size, bold=is_bold)
        cell.font = font
        
        align_kwargs = {}
        if is_centered:
            align_kwargs["horizontal"] = "center"
        if is_vertically_centered:
            align_kwargs["vertical"] = "center"
        if align_kwargs:
            cell.alignment = Alignment(wrap_text=True, **align_kwargs)
            
        if has_borders:
            thin_side = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    def apply_borders_to_range(self, sheet, start_row, start_col, end_row, end_col):
        from openpyxl.styles import Border, Side
        thin_side = Side(border_style="thin", color="000000")
        border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = sheet.cell(row=r, column=c)
                cell.border = border

    def format_total_time(self, total_minutes):
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours} hr {minutes:02d} min"

    def find_or_create_user_row(self, sheet, user_name):
        last_row = sheet.max_row
        for r in range(4, last_row + 1):
            val = sheet.cell(row=r, column=1).value
            if val == user_name:
                return r
        # Create new row
        new_row = last_row + 1
        cell = sheet.cell(row=new_row, column=1, value=user_name)
        self.apply_cell_style(cell, is_bold=True, has_borders=True, is_centered=False, is_vertically_centered=True, font_size=11)
        return new_row

    def write_user_record(self, sheet, user_name, records, time_zone):
        row_num = self.find_or_create_user_row(sheet, user_name)
        
        start_day = int(sheet.cell(row=3, column=2).value) if sheet.cell(row=3, column=2).value else 1
        max_col = sheet.max_column
        end_day = int(sheet.cell(row=3, column=max_col - 1).value) if sheet.cell(row=3, column=max_col - 1).value else start_day
        
        input_format = "%d/%m/%Y, %I:%M:%S %p"
        time_format = "%H:%M"
        
        total_minutes = 0
        day_entries = {}
        
        for record in records:
            if not record.get("timeIn") or not record.get("timeOut"):
                continue
            try:
                # Convert timeIn/timeOut from UTC to local
                time_in_local_str = self.common_service.convert_utc_to_local(record["timeIn"], time_zone)
                time_out_local_str = self.common_service.convert_utc_to_local(record["timeOut"], time_zone)
                
                # Parse local strings
                time_in = datetime.strptime(time_in_local_str, input_format)
                time_out = datetime.strptime(time_out_local_str, input_format)
                
                day_of_month = time_in.day
                duration_minutes = int(round((time_out - time_in).total_seconds() / 60.0))
                
                if duration_minutes < 0:
                    duration_minutes += 24 * 60  # Handle overnight shifts
                    
                total_minutes += duration_minutes
                
                if start_day <= day_of_month <= end_day:
                    if day_of_month not in day_entries:
                        day_entries[day_of_month] = []
                    day_entries[day_of_month].append(f"{time_in.strftime(time_format)} - {time_out.strftime(time_format)}")
            except Exception as e:
                logger.error(f"Error parsing record for excel: {e}")
                
        # Write to cells
        for day in range(start_day, end_day + 1):
            col_idx = day - start_day + 2
            cell = sheet.cell(row=row_num, column=col_idx)
            self.apply_cell_style(cell, is_bold=False, has_borders=True, is_centered=False, is_vertically_centered=True, font_size=11)
            
            if day in day_entries:
                cell.value = "\n".join(day_entries[day])
            else:
                cell.value = "-"
                
        total_cell = sheet.cell(row=row_num, column=end_day - start_day + 3)
        self.apply_cell_style(total_cell, is_bold=False, has_borders=True, is_centered=True, is_vertically_centered=True, font_size=11)
        total_cell.value = self.format_total_time(total_minutes)

    def parse_date_string(self, date_str: str) -> date:
        if date_str is None:
            return None
        if "," in date_str:
            date_str = date_str.split(",")[0].strip()
        return datetime.strptime(date_str.strip(), "%d/%m/%Y").date()

    def is_weekly_off_day(self, date_obj: date, weekly_off: WeeklyOff) -> bool:
        if weekly_off is None:
            return False
            
        day_of_week = date_obj.weekday()  # 0 is Monday, 6 is Sunday
        week_of_month = ((date_obj.day - 1) // 7) + 1  # 1 to 5
        
        if day_of_week == 6:  # SUNDAY
            if weekly_off.sundayAll: return True
            return (week_of_month == 1 and weekly_off.sunday1st) or \
                   (week_of_month == 2 and weekly_off.sunday2nd) or \
                   (week_of_month == 3 and weekly_off.sunday3rd) or \
                   (week_of_month == 4 and weekly_off.sunday4th) or \
                   (week_of_month == 5 and weekly_off.sunday5th)
        elif day_of_week == 0:  # MONDAY
            if weekly_off.mondayAll: return True
            return (week_of_month == 1 and weekly_off.monday1st) or \
                   (week_of_month == 2 and weekly_off.monday2nd) or \
                   (week_of_month == 3 and weekly_off.monday3rd) or \
                   (week_of_month == 4 and weekly_off.monday4th) or \
                   (week_of_month == 5 and weekly_off.monday5th)
        elif day_of_week == 1:  # TUESDAY
            if weekly_off.tuesdayAll: return True
            return (week_of_month == 1 and weekly_off.tuesday1st) or \
                   (week_of_month == 2 and weekly_off.tuesday2nd) or \
                   (week_of_month == 3 and weekly_off.tuesday3rd) or \
                   (week_of_month == 4 and weekly_off.tuesday4th) or \
                   (week_of_month == 5 and weekly_off.tuesday5th)
        elif day_of_week == 2:  # WEDNESDAY
            if weekly_off.wednesdayAll: return True
            return (week_of_month == 1 and weekly_off.wednesday1st) or \
                   (week_of_month == 2 and weekly_off.wednesday2nd) or \
                   (week_of_month == 3 and weekly_off.wednesday3rd) or \
                   (week_of_month == 4 and weekly_off.wednesday4th) or \
                   (week_of_month == 5 and weekly_off.wednesday5th)
        elif day_of_week == 3:  # THURSDAY
            if weekly_off.thursdayAll: return True
            return (week_of_month == 1 and weekly_off.thursday1st) or \
                   (week_of_month == 2 and weekly_off.thursday2nd) or \
                   (week_of_month == 3 and weekly_off.thursday3rd) or \
                   (week_of_month == 4 and weekly_off.thursday4th) or \
                   (week_of_month == 5 and weekly_off.thursday5th)
        elif day_of_week == 4:  # FRIDAY
            if weekly_off.fridayAll: return True
            return (week_of_month == 1 and weekly_off.friday1st) or \
                   (week_of_month == 2 and weekly_off.friday2nd) or \
                   (week_of_month == 3 and weekly_off.friday3rd) or \
                   (week_of_month == 4 and weekly_off.friday4th) or \
                   (week_of_month == 5 and weekly_off.friday5th)
        elif day_of_week == 5:  # SATURDAY
            if weekly_off.saturdayAll: return True
            return (week_of_month == 1 and weekly_off.saturday1st) or \
                   (week_of_month == 2 and weekly_off.saturday2nd) or \
                   (week_of_month == 3 and weekly_off.saturday3rd) or \
                   (week_of_month == 4 and weekly_off.saturday4th) or \
                   (week_of_month == 5 and weekly_off.saturday5th)
        return False

    def dashboard_counts(self, company_id: int) -> dict:
        try:
            res = {}
            tz = pytz.timezone("Asia/Kolkata")
            now = datetime.now(tz)
            start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)

            count_checked_in = UserInOut.objects.filter(
                companyDetails_id=company_id,
                timeIn__range=(start_of_day, end_of_day)
            ).values('user').distinct().count()

            count_checked_out = UserInOut.objects.filter(
                companyDetails_id=company_id,
                timeOut__range=(start_of_day, end_of_day)
            ).values('user').distinct().count()

            total_users = CompanyEmployee.objects.filter(
                companyDetails_id=company_id,
                isActive=1
            ).count()

            res["countCheckedInUsers"] = count_checked_in
            res["countCheckedOutUsers"] = count_checked_out
            res["companyTotalUserCount"] = total_users
            return res
        except Exception as e:
            logger.error(f"Error dashboard_counts: {e}")
            raise Exception(str(e))

    def format_minutes_to_hhmm(self, minutes: int) -> str:
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"

    def get_all_entries_grouped_by_user(self, user_ids: list, start_date_str: str, end_date_str: str, time_zone: str, location_ids: list, department_ids: list, company_id: int) -> dict:
        try:
            zone = pytz.timezone(time_zone)
            if not start_date_str or not end_date_str:
                utc = pytz.utc
                now = datetime.now(utc).date()
                start_local = now.replace(day=1)
                end_local = now
                
                start_zdt = utc.localize(datetime.combine(start_local, time.min))
                end_zdt = utc.localize(datetime.combine(end_local, time.max))
            else:
                start_local = self.parse_date_string(start_date_str)
                end_local = self.parse_date_string(end_date_str)
                
                start_zdt = zone.localize(datetime.combine(start_local, time.min))
                end_zdt = zone.localize(datetime.combine(end_local, time.max))

            start_utc = start_zdt.astimezone(pytz.utc)
            end_utc = end_zdt.astimezone(pytz.utc)

            spec = Q(createdOn__gte=start_utc, createdOn__lte=end_utc)
            if user_ids:
                spec &= Q(user__employeeId__in=user_ids)
            if location_ids:
                spec &= Q(locations__id__in=location_ids)
            if department_ids:
                spec &= Q(user__department__id__in=department_ids)
            if company_id:
                spec &= Q(companyDetails_id=company_id)

            user_in_out_list = UserInOut.objects.filter(spec).select_related(
                'user__companyShift', 'user__department', 'user__weeklyOff', 'user__companyDetails', 'locations'
            ).order_by('id')

            grouped_by_user = {}
            for uio in user_in_out_list:
                if not uio.user:
                    continue
                emp_id = uio.user.employeeId
                if emp_id not in grouped_by_user:
                    grouped_by_user[emp_id] = (uio.user, [])
                grouped_by_user[emp_id][1].append(uio)

            date_range = []
            curr = start_local
            while curr <= end_local:
                date_range.append(curr)
                curr += timedelta(days=1)

            user_groups = []
            for emp_id, (user, entries) in grouped_by_user.items():
                regular_minutes = 0
                break_minutes = user.lunchBreak if user.lunchBreak is not None else 0
                
                if user.companyShift:
                    shift = user.companyShift
                    if shift.totalHours is not None:
                        regular_minutes = round(shift.totalHours * 60)

                entry_by_date = {}
                for uio in entries:
                    local_created = uio.createdOn.astimezone(zone)
                    entry_by_date[local_created.date()] = uio

                holiday_dates = []
                holiday_templates = HolidayTemplates.objects.filter(companyDetails_id=user.companyDetails.id) if user.companyDetails else []
                for template in holiday_templates:
                    details = HolidayTemplateDetails.objects.filter(holidayTemplates_id=template.id)
                    for detail in details:
                        if detail.date:
                            holiday_dates.append(detail.date.strftime("%d/%m/%Y"))

                weekly_off = user.weeklyOff

                present_count = 0
                absent_count = 0
                weekly_off_count = 0
                holiday_count = 0

                data_list = []
                total_gross_minutes = 0
                total_overtime_minutes = 0
                row_index = 1

                for date_val in date_range:
                    uio = entry_by_date.get(date_val)
                    data_item = {}

                    is_holiday = False
                    is_weekly_off = False
                    formatted_current_date = date_val.strftime("%d/%m/%Y")

                    if formatted_current_date in holiday_dates:
                        is_holiday = True

                    if not is_holiday and weekly_off:
                        is_weekly_off = self.is_weekly_off_day(date_val, weekly_off)

                    if is_holiday:
                        holiday_count += 1
                    if is_weekly_off:
                        weekly_off_count += 1

                    has_valid_times = (uio is not None and uio.timeIn is not None and uio.timeOut is not None)
                    status = ""

                    if has_valid_times:
                        diff_sec = (uio.timeOut - uio.timeIn).total_seconds()
                        gross_minutes = int(diff_sec // 60)
                        net_minutes = gross_minutes - break_minutes
                        overtime_minutes = max(0, gross_minutes - regular_minutes - break_minutes)

                        total_gross_minutes += gross_minutes
                        total_overtime_minutes += overtime_minutes

                        data_item["id"] = uio.id
                        data_item["timeIn"] = self.common_service.convert_date_to_string(uio.timeIn, time_zone)
                        data_item["timeOut"] = self.common_service.convert_date_to_string(uio.timeOut, time_zone)
                        data_item["createdOn"] = self.common_service.convert_date_to_string(uio.createdOn, time_zone)
                        data_item["locationId"] = uio.locations.id if uio.locations else None
                        data_item["regular"] = self.format_minutes_to_hhmm(regular_minutes)
                        data_item["breakTime"] = self.format_minutes_to_hhmm(break_minutes)
                        data_item["workHours"] = self.format_minutes_to_hhmm(net_minutes)
                        data_item["overtime"] = self.format_minutes_to_hhmm(overtime_minutes)
                        data_item["totalHours"] = self.format_minutes_to_hhmm(gross_minutes)

                        if is_holiday or is_weekly_off:
                            status = "PW"
                            present_count += 1
                        else:
                            status = "P"
                            present_count += 1
                    else:
                        created_on_dt = zone.localize(datetime.combine(date_val, time.min))
                        data_item["id"] = None
                        data_item["timeIn"] = None
                        data_item["timeOut"] = None
                        data_item["createdOn"] = self.common_service.convert_date_to_string(created_on_dt, time_zone)
                        data_item["locationId"] = None
                        data_item["regular"] = self.format_minutes_to_hhmm(regular_minutes)
                        data_item["breakTime"] = self.format_minutes_to_hhmm(break_minutes)
                        data_item["workHours"] = "00:00"
                        data_item["overtime"] = "00:00"
                        data_item["totalHours"] = "00:00"

                        if is_weekly_off:
                            status = "W"
                        elif is_holiday:
                            status = "H"
                        else:
                            status = "A"
                            absent_count += 1

                    data_item["status"] = status
                    data_item["userName"] = f"{user.firstName} {user.lastName}"
                    data_item["rowId"] = row_index
                    row_index += 1
                    data_list.append(data_item)

                user_group = {
                    "id": user.employeeId,
                    "username": f"{user.firstName} {user.lastName}",
                    "presentCount": present_count,
                    "absentCount": absent_count,
                    "weeklyOffCount": weekly_off_count,
                    "holidayCount": holiday_count,
                    "department": user.department.departmentName if user.department else "",
                    "data": data_list,
                    "totalHours": self.format_minutes_to_hhmm(total_gross_minutes),
                    "totalOvertime": self.format_minutes_to_hhmm(total_overtime_minutes)
                }
                user_groups.append(user_group)

            return {"users": user_groups}
        except Exception as e:
            logger.error(f"Error get_all_entries_grouped_by_user: {e}")
            raise Exception(str(e))

    def get_all_entries_by_user_id(self, user_ids: list, start_date_str: str, end_date_str: str, time_zone: str, location_ids: list, department_ids: list, company_id: int) -> list:
        try:
            zone = pytz.timezone(time_zone)
            if not start_date_str or not end_date_str:
                utc = pytz.utc
                now = datetime.now(utc).date()
                start_local = now.replace(day=1)
                end_local = now
                
                start_zdt = utc.localize(datetime.combine(start_local, time.min))
                end_zdt = utc.localize(datetime.combine(end_local, time.max))
            else:
                start_local = self.parse_date_string(start_date_str)
                end_local = self.parse_date_string(end_date_str)
                
                start_zdt = zone.localize(datetime.combine(start_local, time.min))
                end_zdt = zone.localize(datetime.combine(end_local, time.max))

            start_utc = start_zdt.astimezone(pytz.utc)
            end_utc = end_zdt.astimezone(pytz.utc)

            spec = Q(createdOn__gte=start_utc, createdOn__lte=end_utc)
            if user_ids:
                spec &= Q(user__employeeId__in=user_ids)
            if location_ids:
                spec &= Q(locations__id__in=location_ids)
            if department_ids:
                spec &= Q(user__department__id__in=department_ids)
            if company_id:
                spec &= Q(companyDetails_id=company_id)

            user_in_out_list = UserInOut.objects.filter(spec).select_related(
                'user__companyShift', 'user__department', 'user__weeklyOff', 'user__companyDetails', 'locations'
            ).order_by('id')

            dto_list = []
            for uio in user_in_out_list:
                user = uio.user
                if not user:
                    continue
                dto = {
                    "id": uio.id,
                    "userName": f"{user.firstName} {user.lastName}",
                    "hourlyRate": user.hourlyRate,
                    "firstName": user.firstName,
                    "lastName": user.lastName,
                    "createdOn": self.common_service.convert_date_to_string(uio.createdOn, time_zone),
                    "timeIn": self.common_service.convert_date_to_string(uio.timeIn, time_zone) if uio.timeIn else None,
                    "timeOut": self.common_service.convert_date_to_string(uio.timeOut, time_zone) if uio.timeOut else None,
                    "locationId": uio.locations.id if uio.locations else None,
                    "userId": user.employeeId,
                    "isSalaryGenerate": uio.isSalaryGenerate,
                    "companyId": uio.companyDetails.id if uio.companyDetails else None
                }

                company_shift_dto = None
                if user.companyShift:
                    shift = user.companyShift
                    company_shift_dto = {
                        "id": shift.id,
                        "shiftName": shift.shiftName,
                        "startTime": self.common_service.convert_date_to_string(shift.startTime, time_zone) if shift.startTime else None,
                        "endTime": self.common_service.convert_date_to_string(shift.endTime, time_zone) if shift.endTime else None,
                        "totalHours": shift.totalHours,
                        "companyId": shift.companyDetails.id if shift.companyDetails else None
                    }
                dto["companyShiftDto"] = company_shift_dto

                regular_minutes = 0
                break_minutes = user.lunchBreak if user.lunchBreak is not None else 0
                if user.companyShift and user.companyShift.totalHours is not None:
                    regular_minutes = round(user.companyShift.totalHours * 60)

                if uio.timeIn and uio.timeOut:
                    diff_sec = (uio.timeOut - uio.timeIn).total_seconds()
                    gross_minutes = int(diff_sec // 60)
                    work_minutes = gross_minutes - break_minutes
                    overtime_minutes = max(0, gross_minutes - regular_minutes - break_minutes)

                    dto["regular"] = self.format_minutes_to_hhmm(regular_minutes)
                    dto["breakTime"] = self.format_minutes_to_hhmm(break_minutes)
                    dto["workHours"] = self.format_minutes_to_hhmm(work_minutes)
                    dto["overtime"] = self.format_minutes_to_hhmm(overtime_minutes)
                    dto["totalHours"] = self.format_minutes_to_hhmm(gross_minutes)
                    dto["status"] = "P"
                    dto["department"] = user.department.departmentName if user.department else ""
                else:
                    dto["regular"] = self.format_minutes_to_hhmm(regular_minutes)
                    dto["breakTime"] = self.format_minutes_to_hhmm(break_minutes)
                    dto["workHours"] = "00:00"
                    dto["overtime"] = "00:00"
                    dto["totalHours"] = "00:00"
                    dto["status"] = "A"
                    dto["department"] = user.department.departmentName if user.department else ""

                dto["status"] = "P" if uio.timeIn else "A"
                dto_list.append(dto)

            return dto_list
        except Exception as e:
            logger.error(f"Error get_all_entries_by_user_id: {e}")
            raise Exception(str(e))

    def get_user_last_inout(self, id: int) -> dict:
        try:
            user_in_out = UserInOut.objects.filter(user_id=id).order_by('-id').first()
            if user_in_out:
                dto = {
                    "id": user_in_out.id,
                    "userId": user_in_out.user.employeeId if user_in_out.user else None,
                    "timeIn": self.common_service.convert_date_to_string(user_in_out.timeIn, "Asia/Calcutta") if user_in_out.timeIn else None,
                    "timeOut": self.common_service.convert_date_to_string(user_in_out.timeOut, "Asia/Calcutta") if user_in_out.timeOut else None,
                    "locationId": user_in_out.locations.id if user_in_out.locations else None,
                    "createdOn": self.common_service.convert_date_to_string(user_in_out.createdOn, "Asia/Calcutta") if user_in_out.createdOn else None,
                    "companyId": user_in_out.companyDetails.id if user_in_out.companyDetails else None,
                    "isSalaryGenerate": user_in_out.isSalaryGenerate
                }
                return dto
            return None
        except Exception as e:
            logger.error(f"Error get_user_last_inout: {e}")
            raise Exception(str(e))

    def get_user_inout(self, id: int) -> dict:
        try:
            user_in_out = UserInOut.objects.get(id=id)
            dto = {
                "id": user_in_out.id,
                "userId": user_in_out.user.employeeId if user_in_out.user else None,
                "timeIn": self.common_service.convert_date_to_string(user_in_out.timeIn, "Asia/Calcutta") if user_in_out.timeIn else None,
                "timeOut": self.common_service.convert_date_to_string(user_in_out.timeOut, "Asia/Calcutta") if user_in_out.timeOut else None,
                "locationId": user_in_out.locations.id if user_in_out.locations else None
            }
            return dto
        except Exception as e:
            logger.error(f"Error get_user_inout: {e}")
            raise Exception(str(e))

    def create_user_inout(self, user_id: int, location_id: int = None, company_id: int = None, time_in: datetime = None) -> dict:
        try:
            if time_in is None:
                time_in = django.utils.timezone.now()
            elif time_in.tzinfo is None:
                time_in = pytz.utc.localize(time_in)

            employee = CompanyEmployee.objects.get(employeeId=user_id)
            company = CompanyDetails.objects.get(id=company_id) if company_id else employee.companyDetails

            user_in_out = UserInOut()
            user_in_out.user = employee
            user_in_out.companyDetails = company
            user_in_out.timeIn = time_in
            user_in_out.createdOn = time_in
            user_in_out.isSalaryGenerate = 0

            if location_id and location_id > 0:
                user_in_out.locations = Locations.objects.get(id=location_id)

            user_in_out.save()
            return {"id": user_in_out.id}
        except Exception as e:
            logger.error(f"Error create_user_inout: {e}")
            raise Exception(str(e))

    def handle_timeout_update(self, employee: CompanyEmployee, existing_record: UserInOut, time_out: datetime, location_id: int, company_id: int) -> bool:
        auto_time_in_after = employee.companyDetails.autoTimeInAfterHours if (employee.companyDetails and hasattr(employee.companyDetails, 'autoTimeInAfterHours')) else None
        
        if time_out.tzinfo is None:
            time_out = pytz.utc.localize(time_out)

        if not auto_time_in_after or auto_time_in_after.strip() == "":
            existing_record.timeOut = time_out
            existing_record.save()
            return True

        ist_zone = pytz.timezone("Asia/Kolkata")
        time_in_ist = existing_record.timeIn.astimezone(ist_zone)
        time_out_ist = time_out.astimezone(ist_zone)

        try:
            parts = auto_time_in_after.split(":")
            limit_hours = int(parts[0])
            limit_minutes = int(parts[1])
        except Exception:
            limit_hours = 0
            limit_minutes = 0

        allowed_limit = timedelta(hours=limit_hours, minutes=limit_minutes)
        session_duration = time_out_ist - time_in_ist

        if session_duration > allowed_limit:
            next_day_time_in = time_out + timedelta(days=1)
            self.create_user_inout(employee.employeeId, location_id, company_id, next_day_time_in)
            return True
        else:
            existing_record.timeOut = time_out
            existing_record.save()
            return True

    def update_user_inout_by_id(self, id: int, user_id: int) -> None:
        try:
            user_in_out = UserInOut.objects.get(id=id)
            employee = CompanyEmployee.objects.get(employeeId=user_id)
            now = django.utils.timezone.now()
            location_id = user_in_out.locations.id if user_in_out.locations else None
            company_id = employee.companyDetails.id if employee.companyDetails else None
            self.handle_timeout_update(employee, user_in_out, now, location_id, company_id)
        except Exception as e:
            logger.error(f"Error update_user_inout_by_id: {e}")
            raise Exception(str(e))

    def update_user_inout_by_dto(self, dto: dict) -> dict:
        try:
            user_in_out = UserInOut.objects.get(id=dto["id"])
            employee = CompanyEmployee.objects.get(employeeId=dto["userId"])

            if not dto.get("timeOut"):
                raise Exception("TimeOut is required for update")

            time_out = self.common_service.convert_string_to_date(dto["timeOut"])
            location_id = dto.get("locationId")
            company_id = employee.companyDetails.id if employee.companyDetails else None

            self.handle_timeout_update(employee, user_in_out, time_out, location_id, company_id)
            return dto
        except Exception as e:
            logger.error(f"Error update_user_inout_by_dto: {e}")
            raise Exception(str(e))

    def click_in_out(self, user_id: int, location_id: int = None, company_id: int = None) -> str:
        try:
            employee = CompanyEmployee.objects.get(employeeId=user_id)
            existing = UserInOut.objects.filter(user_id=user_id, timeOut__isnull=True).order_by('-id').first()

            if existing:
                self.update_user_inout_by_id(existing.id, user_id)
                username = employee.userName if employee.userName else f"{employee.firstName} {employee.lastName}"
                return f"updated:{username}"
            else:
                self.create_user_inout(user_id, location_id, company_id, django.utils.timezone.now())
                username = employee.userName if employee.userName else f"{employee.firstName} {employee.lastName}"
                return f"created:{username}"
        except Exception as e:
            logger.error(f"Error click_in_out: {e}")
            raise Exception(str(e))

    def parse_any_date(self, s: str) -> datetime:
        if not s or not s.strip():
            return None
        s = s.strip()

        if "T" in s:
            if s.endswith('Z'):
                s = s[:-1] + '+00:00'
            try:
                parts = s.split('.')
                if len(parts) > 1 and '+' in parts[1]:
                    subparts = parts[1].split('+')
                    if len(subparts[0]) > 6:
                        parts[1] = subparts[0][:6] + '+' + subparts[1]
                        s = parts[0] + '.' + parts[1]
                return datetime.fromisoformat(s)
            except Exception:
                pass

        formats = [
            "%d/%m/%Y, %I:%M:%S %p",
            "%d/%m/%Y, %H:%M:%S",
            "%d/%m/%Y",
            "%Y-%m-%d"
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(s, fmt)
                return pytz.utc.localize(dt)
            except ValueError:
                continue

        raise Exception(f"Invalid date format: {s}")

    def add_clock_in_out(self, dto: dict) -> dict:
        try:
            if dto.get("id") is not None:
                user_in_out = UserInOut.objects.get(id=dto["id"])
                employee = CompanyEmployee.objects.get(employeeId=dto["userId"])

                if dto.get("timeOut") is not None:
                    time_out = self.parse_any_date(dto["timeOut"])
                    self.handle_timeout_update(employee, user_in_out, time_out, dto.get("locationId"), dto.get("companyId"))

                if dto.get("timeIn") is not None:
                    user_in_out.timeIn = self.parse_any_date(dto["timeIn"])

                user_in_out.save()
                return dto
            else:
                user_in_out = UserInOut()
                employee = CompanyEmployee.objects.get(employeeId=dto["userId"])
                company = CompanyDetails.objects.get(id=dto["companyId"])

                user_in_out.isSalaryGenerate = 0
                user_in_out.user = employee
                user_in_out.companyDetails = company

                if dto.get("createdOn") is None:
                    user_in_out.createdOn = django.utils.timezone.now()
                else:
                    user_in_out.createdOn = self.parse_any_date(dto["createdOn"])

                if dto.get("timeIn") is not None:
                    user_in_out.timeIn = self.parse_any_date(dto["timeIn"])
                else:
                    raise Exception("Clock In is required")

                if dto.get("timeOut") is not None:
                    user_in_out.timeOut = self.parse_any_date(dto["timeOut"])
                else:
                    user_in_out.timeOut = None

                user_in_out.save()
                dto["id"] = user_in_out.id
                return dto
        except Exception as e:
            logger.error(f"Error add_clock_in_out: {e}")
            raise Exception(str(e))

    def get_today_entries_by_user_id(self, user_id: int) -> list:
        try:
            tz = pytz.timezone("Asia/Kolkata")
            now = datetime.now(tz)
            start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)

            entries = UserInOut.objects.filter(user_id=user_id, createdOn__range=(start_of_day, end_of_day))

            dto_list = []
            for entry in entries:
                dto_list.append({
                    "id": entry.id,
                    "timeIn": self.common_service.convert_date_to_string(entry.timeIn, "Asia/Calcutta") if entry.timeIn else None,
                    "timeOut": self.common_service.convert_date_to_string(entry.timeOut, "Asia/Calcutta") if entry.timeOut else None,
                    "userId": entry.user.employeeId if entry.user else None
                })
            return dto_list
        except Exception as e:
            logger.error(f"Error get_today_entries_by_user_id: {e}")
            raise Exception(str(e))

    def get_time_inout_report(self, user_ids: list, start_date_str: str, end_date_str: str, time_zone: str, company_id: int) -> dict:
        try:
            if not start_date_str or not end_date_str:
                now = datetime.now(pytz.utc)
                start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                if now.month == 12:
                    next_month = now.replace(year=now.year + 1, month=1, day=1, hour=23, minute=59, second=59, microsecond=999999)
                else:
                    next_month = now.replace(month=now.month + 1, day=1, hour=23, minute=59, second=59, microsecond=999999)
                end = next_month - timedelta(days=1)
            else:
                start = self.common_service.convert_local_to_utc(start_date_str, time_zone, False)
                end = self.common_service.convert_local_to_utc(end_date_str, time_zone, True)

            if user_ids:
                users = CompanyEmployee.objects.filter(employeeId__in=user_ids)
            else:
                if company_id:
                    users = CompanyEmployee.objects.filter(companyDetails_id=company_id)
                else:
                    users = CompanyEmployee.objects.all()

            user_map = {user.employeeId: f"{user.firstName} {user.lastName}" for user in users}

            spec = Q(createdOn__gte=start, createdOn__lte=end)
            if user_ids:
                spec &= Q(user__employeeId__in=user_ids)
            if company_id:
                spec &= Q(companyDetails_id=company_id)

            user_in_out_records = UserInOut.objects.filter(spec).select_related('user')
            user_in_out_map = {}
            for record in user_in_out_records:
                if not record.user:
                    continue
                emp_id = record.user.employeeId
                if emp_id not in user_in_out_map:
                    user_in_out_map[emp_id] = []
                user_in_out_map[emp_id].append(record)

            response_list = []
            for current_user_id, user_name in user_map.items():
                records = user_in_out_map.get(current_user_id, [])
                monthly_records = {}
                for record in records:
                    month = record.createdOn.month
                    time_in_str = self.common_service.convert_date_to_string(record.timeIn, time_zone) if record.timeIn else ""
                    time_out_str = self.common_service.convert_date_to_string(record.timeOut, time_zone) if record.timeOut else ""
                    
                    day_record = {
                        "records": [
                            {
                                "timeIn": time_in_str,
                                "timeOut": time_out_str
                            }
                        ]
                    }
                    if month not in monthly_records:
                        monthly_records[month] = []
                    monthly_records[month].append(day_record)

                month_data = []
                for m, data_list in monthly_records.items():
                    month_data.append({
                        "month": m,
                        "data": data_list
                    })
                response_list.append({
                    "username": user_name,
                    "records": month_data
                })

            return {"data": response_list}
        except Exception as e:
            logger.error(f"Error get_time_inout_report: {e}")
            raise Exception(str(e))

    def generate_excel_report(self, data: dict, start_date_str: str, end_date_str: str, time_zone: str):
        import openpyxl
        
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        month_format = "%B %Y"
        input_date_format = "%m/%d/%Y"

        start_date = None
        end_date = None

        try:
            if start_date_str:
                start_date = datetime.strptime(start_date_str.split(",")[0].strip(), input_date_format).date()
            if end_date_str:
                end_date = datetime.strptime(end_date_str.split(",")[0].strip(), input_date_format).date()
        except Exception as e:
            logger.error(f"Invalid start/end date format in excel generation: {e}")

        if not start_date or not end_date:
            now = datetime.now()
            start_date = now.replace(day=1).date()
            if now.month == 12:
                next_month = now.replace(year=now.year + 1, month=1, day=1)
            else:
                next_month = now.replace(month=now.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).date()

        month_keys = []
        curr = start_date
        while curr <= end_date:
            m_key = curr.strftime(month_format)
            if m_key not in month_keys:
                month_keys.append(m_key)
            # Move safely to next month: add 28 days first, then loop to find next month start
            curr_month = curr.month
            while curr.month == curr_month:
                curr += timedelta(days=1)

        month_sheets = {}
        for m_key in month_keys:
            sheet_title = m_key.replace("/", "-")
            sheet = workbook.create_sheet(title=sheet_title)

            try:
                sheet_month_date = datetime.strptime(m_key, month_format).date()
            except Exception:
                sheet_month_date = start_date

            sheet_start_date = sheet_month_date.replace(day=1)
            if sheet_start_date.month == 12:
                sheet_end_date = (sheet_start_date.replace(year=sheet_start_date.year + 1, month=1) - timedelta(days=1))
            else:
                sheet_end_date = (sheet_start_date.replace(month=sheet_start_date.month + 1) - timedelta(days=1))

            start_day = 1
            end_day = sheet_end_date.day

            if start_date.month == sheet_start_date.month and start_date.year == sheet_start_date.year:
                start_day = start_date.day
            if end_date.month == sheet_end_date.month and end_date.year == sheet_end_date.year:
                end_day = end_date.day

            total_columns = end_day - start_day + 2

            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns + 1)
            title_cell = sheet.cell(row=1, column=1, value="In-Out Report")
            self.apply_cell_style(title_cell, is_bold=True, has_borders=True, is_centered=True, is_vertically_centered=True, font_size=14)
            self.apply_borders_to_range(sheet, 1, 1, 1, total_columns + 1)

            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns + 1)
            month_cell = sheet.cell(row=2, column=1, value=m_key)
            self.apply_cell_style(month_cell, is_bold=True, has_borders=True, is_centered=True, is_vertically_centered=True, font_size=14)
            self.apply_borders_to_range(sheet, 2, 1, 2, total_columns + 1)

            uname_cell = sheet.cell(row=3, column=1, value="User Name")
            self.apply_cell_style(uname_cell, is_bold=True, has_borders=True, is_centered=True, is_vertically_centered=True, font_size=12)

            for d in range(start_day, end_day + 1):
                cell = sheet.cell(row=3, column=d - start_day + 2, value=d)
                self.apply_cell_style(cell, is_bold=True, has_borders=True, is_centered=True, is_vertically_centered=True, font_size=12)

            tot_cell = sheet.cell(row=3, column=total_columns + 1, value="Total Hours")
            self.apply_cell_style(tot_cell, is_bold=True, has_borders=True, is_centered=True, is_vertically_centered=True, font_size=12)

            month_sheets[m_key] = sheet

        user_data_list = data.get("data", [])
        for user in user_data_list:
            user_name = user.get("username")
            records = user.get("records", [])

            if not records:
                for m_key in month_sheets.keys():
                    sheet = month_sheets[m_key]
                    self.write_user_record(sheet, user_name, [], time_zone)
                continue

            for m_key in month_sheets.keys():
                sheet = month_sheets[m_key]
                filtered_records = []
                try:
                    current_month_date = datetime.strptime(m_key, month_format).date()
                    current_month = current_month_date.month

                    for record_group in records:
                        record_month = record_group.get("month")
                        if record_month == current_month:
                            data_list = record_group.get("data", [])
                            for data_entry in data_list:
                                time_records = data_entry.get("records", [])
                                filtered_records.extend(time_records)
                    self.write_user_record(sheet, user_name, filtered_records, time_zone)
                except Exception as e:
                    logger.error(f"Error writing user record for month key {m_key}: {e}")

        return workbook

    def delete_user_inout(self, id: int) -> None:
        try:
            user_in_out = UserInOut.objects.get(id=id)
            user_in_out.delete()
        except Exception as e:
            logger.error(f"Error delete_user_inout: {e}")
            raise Exception(str(e))

    def add_bulk_clock_in_out(self, bulk_dto: dict) -> None:
        try:
            user_ids = bulk_dto.get("userId")
            if not user_ids:
                raise Exception("User ID list cannot be empty")
            if not bulk_dto.get("startDate") or not bulk_dto.get("endDate"):
                raise Exception("Start date and End date are required")

            start_dt = self.parse_any_date(bulk_dto.get("startDate"))
            end_dt = self.parse_any_date(bulk_dto.get("endDate"))

            if not start_dt or not end_dt:
                raise Exception("Start date and End date are invalid")

            start_local = start_dt.astimezone(pytz.utc).date()
            end_local = end_dt.astimezone(pytz.utc).date()

            if start_local > end_local:
                raise Exception("Start date cannot be after End date")

            time_in_dt = self.parse_any_date(bulk_dto.get("timeIn"))
            time_out_dt = self.parse_any_date(bulk_dto.get("timeOut")) if bulk_dto.get("timeOut") else None

            if not time_in_dt:
                raise Exception("Time In is required")

            base_time_in = time_in_dt.astimezone(pytz.utc)
            base_time_out = time_out_dt.astimezone(pytz.utc) if time_out_dt else None
            duration = base_time_out - base_time_in if base_time_out else None

            try:
                company_details = CompanyDetails.objects.get(id=bulk_dto.get("companyId"))
            except CompanyDetails.DoesNotExist:
                raise Exception("Company not found")

            for user_id in user_ids:
                try:
                    company_employee = CompanyEmployee.objects.get(employeeId=user_id)
                except CompanyEmployee.DoesNotExist:
                    raise Exception(f"Employee not found for id: {user_id}")

                weekly_off = company_employee.weeklyOff

                holiday_dates = []
                if company_employee.holidayTemplates:
                    details = HolidayTemplateDetails.objects.filter(holidayTemplates_id=company_employee.holidayTemplates.id)
                    for detail in details:
                        if detail.date:
                            holiday_dates.append(detail.date.strftime("%d/%m/%Y"))

                curr_date = start_local
                while curr_date <= end_local:
                    formatted_date = curr_date.strftime("%d/%m/%Y")
                    if holiday_dates and formatted_date in holiday_dates:
                        curr_date += timedelta(days=1)
                        continue

                    if weekly_off:
                        if self.is_weekly_off_day(curr_date, weekly_off):
                            curr_date += timedelta(days=1)
                            continue

                    day_start = pytz.utc.localize(datetime.combine(curr_date, time.min))
                    day_end = pytz.utc.localize(datetime.combine(curr_date, time.max))

                    existing_records = UserInOut.objects.filter(
                        user_id=company_employee.employeeId,
                        createdOn__gte=day_start,
                        createdOn__lte=day_end
                    )

                    if not existing_records.exists():
                        user_in_out = UserInOut()
                        user_in_out.user = company_employee
                        user_in_out.companyDetails = company_details
                        user_in_out.isSalaryGenerate = 0

                        new_time_in = base_time_in.replace(year=curr_date.year, month=curr_date.month, day=curr_date.day)
                        new_time_out = new_time_in + duration if duration else None

                        user_in_out.timeIn = new_time_in
                        user_in_out.timeOut = new_time_out
                        user_in_out.createdOn = new_time_in
                        user_in_out.save()

                    curr_date += timedelta(days=1)
        except Exception as e:
            logger.error(f"Error in add_bulk_clock_in_out: {e}")
            raise Exception(str(e))
